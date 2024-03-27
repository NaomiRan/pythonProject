import openpyxl as xl
from openpyxl.chart import BarChart, Series, Reference

# create a new worksheet
# wb = Workbook()
# ws = wb.active


# define a reusable function
def process_workbook(filename):
       wb=xl.load_workbook(filename)
       sheet=wb['Sheet1']
       for row in range(2,sheet.max_row+1):
          cell=sheet.cell(row,3)
          corrected_price=cell.value * 0.9
          corrected_price_cell=sheet.cell(row,4)
          corrected_price_cell.value=corrected_price

       values=Reference(sheet,
                        min_row=2,
                        max_row=sheet.max_row,
                        min_col=4,
                        max_col=4)
       chart=BarChart()
       chart.title='transaction'
       chart.x_axis='transaction_id'
       chart.y_axis='price'
       chart.legend= None
       chart.add_data(values)
       sheet.add_chart(chart,'e2')

       wb.save(filename)


process_workbook('transactions.xlsx')